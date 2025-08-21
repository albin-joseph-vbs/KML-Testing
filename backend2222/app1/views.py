
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
#(2) Views for the User Login Views for the User Login
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
# #Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError
from .serializers import LoginSerializer

from django.shortcuts import render

def dojo_app(request):
    return render (request,'index.html')

class LoginAPIView(APIView):

    """
    User Login API View
    """
    def post(self, request):
        print(request.data)
        serializer = LoginSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Log the user in
            login(request, user)
            
            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            print(user.role)
            
            # Return user info & tokens
            return Response({
                'message': 'Login successful',
                'access_token': access_token,
                'refresh_token': str(refresh),
                'user': {
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'employeeid': user.employeeid,
                    'role': user.role,
                    'hq': user.hq,
                    'factory': user.factory,
                    'department': user.department,
                    'status': user.status
                }
            }, status=status.HTTP_200_OK)
        
        # Extract error message
        error_message = "Authentication failed"
        if serializer.errors:
            for field, errors in serializer.errors.items():
                if errors:
                    if isinstance(errors, list) and errors:
                        error_message = errors[0]
                    else:
                        error_message = str(errors)
                    break
        
        return Response({
            'error': True,
            'message': error_message
        }, status=status.HTTP_400_BAD_REQUEST)
    





#(1) Views for the User Register

from django.db import IntegrityError
from django.shortcuts import render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import User

class RegisterView(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        user_data = request.data
        serializer = self.serializer_class(data=user_data)

        try:
            serializer.is_valid(raise_exception=True)

            # Check if user already exists by email
            if User.objects.filter(email=user_data.get("email")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"email": "This email is already registered."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee ID already exists
            if User.objects.filter(employeeid=user_data.get("employeeid")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"employeeid": "This employee ID is already in use."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save the user
            serializer.save()

            return Response({
                "message": "User registered successfully!"
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            # Handle specific validation errors
            return Response({
                "message": "Validation failed",
                "errors": e.detail
            }, status=status.HTTP_400_BAD_REQUEST)

        except IntegrityError:
            # Handle database integrity errors (like duplicate entries)
            return Response({
                "message": "Database error",
                "errors": {"detail": "Duplicate entry or constraint violation."}
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # Handle unexpected errors
            return Response({
                "message": "Unexpected error occurred",
                "errors": {"detail": str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#(3) Views for the User Logout

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LogoutSerializer

class LogoutAPIView(APIView):
    """
    User Logout API View
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            refresh_token = serializer.validated_data["refresh_token"]

            try:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Blacklist the refresh token
                return Response({"message": "Logout successful"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


















from django.shortcuts import render

# Create your views here.
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import EmployeeMaster
from .serializers import EmployeeSerializer

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMaster.objects.all()
    serializer_class = EmployeeSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            # Print validation errors to console for debugging
            print("Validation errors:", serializer.errors)
            # Return detailed error response
            return Response({
                'status': 'error',
                'errors': serializer.errors,
                'message': 'Invalid data provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response({
                'status': 'success',
                'data': serializer.data,
                'message': 'Employee created successfully'
            }, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            # Print exception to console
            print("Exception occurred:", str(e))
            return Response({
                'status': 'error',
                'message': str(e),
                'details': 'An error occurred while creating employee'
            }, status=status.HTTP_400_BAD_REQUEST)


from rest_framework import viewsets
from .models import (
     Station, OperatorSkill,
    TrainingTopic, OperatorTraining, MonthlyAssignment
)
from .serializers import (
     StationSerializer, OperatorSkillSerializer,
    TrainingTopicSerializer, OperatorTrainingSerializer, MonthlyAssignmentSerializer
)


# class OperatorViewSet(viewsets.ModelViewSet):
#     queryset = Operator.objects.all()
#     serializer_class = OperatorSerializer


class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer


class OperatorSkillViewSet(viewsets.ModelViewSet):
    queryset = OperatorSkill.objects.all()
    serializer_class = OperatorSkillSerializer


class TrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = TrainingTopic.objects.all()
    serializer_class = TrainingTopicSerializer


class OperatorTrainingViewSet(viewsets.ModelViewSet):
    queryset = OperatorTraining.objects.all()
    serializer_class = OperatorTrainingSerializer


class MonthlyAssignmentViewSet(viewsets.ModelViewSet):
    queryset = MonthlyAssignment.objects.all()
    serializer_class = MonthlyAssignmentSerializer









from rest_framework import viewsets
from .models import HQ, Factory, Department, Line, Level
from .serializers import HQSerializer, FactorySerializer, DepartmentSerializer, LineSerializer, LevelSerializer

class HQViewSet(viewsets.ModelViewSet):
    queryset = HQ.objects.all()
    serializer_class = HQSerializer

class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer

class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer





from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelTracking, EmployeeMaster
from .serializers import OperatorLevelTrackingSerializer
from datetime import date, timedelta

@api_view(['GET'])
def get_today_milestones(request):
    today = date.today()
    milestone_rules = OperatorLevelTracking.objects.all()
    milestone_data = []

    for rule in milestone_rules:
        expected_join_date = today - timedelta(days=rule.day)
        matched_employees = EmployeeMaster.objects.filter(joining_date=expected_join_date)

        for employee in matched_employees:
            milestone_data.append({
                "operator_name": employee.name,
                "level_name": rule.level.name,
                "day": rule.day,
                "milestone_date": today,
                "message": f"{employee.name} has completed milestone: {rule.level.name} on Day {rule.day}"
            })

    return Response({
        "date": str(today),
        "milestones": milestone_data
    })






from datetime import date, timedelta
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelEmailTracking, TrackingEmail, EmployeeMaster
from .utils import send_milestone_email

@api_view(['GET'])
def get_today_email_milestones(request):
    today = date.today()
    rules = OperatorLevelEmailTracking.objects.all()
    email_logs = []

    for rule in rules:
        milestone_day = rule.day
        expected_date = today - timedelta(days=milestone_day)

        # Get all employees who joined exactly 'day' days ago
        matching_employees = EmployeeMaster.objects.filter(joining_date=expected_date)

        recipient_list = [e.email for e in rule.emails.all() if e.email]

        for employee in matching_employees:
            if recipient_list:
                subject = "Milestone Alert"
                message = f"{employee.name} has reached milestone: {rule.level.name} on Day {rule.day}."
                send_milestone_email(subject, message, recipient_list)
                email_logs.append({
                    "employee": employee.name,
                    "joined_on": str(employee.joining_date),
                    "level": rule.level.name,
                    "day": rule.day,
                    "recipients": recipient_list,
                    "status": "Email sent"
                })

    return Response({
        "message": "Milestone emails sent to matching employees.",
        "email_logs": email_logs
    })









from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import Machine, MachineAllocation
from .serializers import MachineSerializer, MachineAllocationSerializer

class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all()
    serializer_class = MachineSerializer


from rest_framework import viewsets, status
from rest_framework.response import Response
from django.core.mail import send_mail
from .models import MachineAllocation, MachineAllocationTrackingEmail
from .serializers import MachineAllocationSerializer

class MachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            self.perform_create(serializer)

            # Send email to all tracking emails
            allocation = serializer.instance
            subject = "Machine Allocation Approval Request"
            message = (
                f"Machine '{allocation.machine.name}' has been allocated to "
                f"'{allocation.employee.name}'.\n\n"
                "Please review and approve this allocation request."
            )
            from_email = None  # Will use DEFAULT_FROM_EMAIL
            recipient_list = list(MachineAllocationTrackingEmail.objects.values_list('email', flat=True))

            if recipient_list:
                send_mail(subject, message, from_email, recipient_list)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # Ensure only 'approval_status' is being updated
        if set(serializer.validated_data.keys()) != {'approval_status'}:
            return Response(
                {'error': 'Only approval_status can be updated.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        self.perform_update(serializer)
        return Response(serializer.data)







from rest_framework import viewsets
from .models import SkillTraining
from .serializers import SkillTrainingSerializer

class SkillTrainingViewSet(viewsets.ModelViewSet):
    queryset = SkillTraining.objects.all()
    serializer_class = SkillTrainingSerializer






from rest_framework import viewsets
from .models import SubTopic
from .serializers import SubTopicSerializer

class SubTopicViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer



from rest_framework import viewsets
from .models import SubTopic
from .serializers import SubTopicDaySerializer

class SubTopicDayViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicDaySerializer




from rest_framework import viewsets
from .models import SubTopicContent
from .serializers import SubTopicContentSerializer

class SubTopicContentViewSet(viewsets.ModelViewSet):
    queryset = SubTopicContent.objects.all()
    serializer_class = SubTopicContentSerializer





from rest_framework import viewsets
from .models import Days
from .serializers import DaysSerializer

class DaysViewSet(viewsets.ModelViewSet):
    queryset = Days.objects.all()
    serializer_class = DaysSerializer







from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer

class TrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

from rest_framework.views import APIView
class TrainingContentCreateView(APIView):
    def post(self, request):
        serializer = TrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    




from rest_framework import viewsets
from .models import LevelTwoProduction, LevelTwoLine,LevelTwoSubStation
from .serializers import LevelTwoProductionSerializer, LevelTwoLineSerializer,LevelTwoSubStationSerializer

class LevelTwoProductionViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoProduction.objects.all()
    serializer_class = LevelTwoProductionSerializer


class LevelTwoLineViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoLine.objects.all()
    serializer_class = LevelTwoLineSerializer



class LevelTwoSubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSubStation.objects.all()
    serializer_class = LevelTwoSubStationSerializer

    def get_queryset(self):
        line_id = self.request.query_params.get('line_id')
        if line_id:
            return LevelTwoSubStation.objects.filter(line_id=line_id)
        return LevelTwoSubStation.objects.all()


from rest_framework import viewsets
from .models import (
    LevelTwoTraineeInfo,
    LevelTwoTrainingTopic,
    LevelTwoOJTDay,
    LevelTwoOJTScore,
)
from .serializers import (
    LevelTwoTraineeInfoSerializer,
    LevelTwoTrainingTopicSerializer,
    LevelTwoOJTDaySerializer,
    LevelTwoOJTScoreSerializer,
)

class LevelTwoTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTraineeInfo.objects.all()
    serializer_class = LevelTwoTraineeInfoSerializer


class LevelTwoTrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTrainingTopic.objects.all()
    serializer_class = LevelTwoTrainingTopicSerializer


class LevelTwoOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoOJTDay.objects.all()
    serializer_class = LevelTwoOJTDaySerializer


from rest_framework import viewsets
from .models import LevelTwoOJTScore
from .serializers import LevelTwoOJTScoreSerializer
from .utils import check_and_update_operator_skill  # Make sure to import this

class LevelTwoOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoOJTScore.objects.all()
    serializer_class = LevelTwoOJTScoreSerializer

    def perform_create(self, serializer):
        # Save the new OJT score
        instance = serializer.save()

        # Update training status
        instance.trainee.calculate_and_save_training_status()

        # ✅ Call the check_and_update_operator_skill function
        if instance.trainee.traineeId:
            check_and_update_operator_skill(instance.trainee.traineeId)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster

class EmployeeNameByCodeAPIView(APIView):
    def get(self, request, pay_code):
        try:
            employee = EmployeeMaster.objects.get(pay_code=pay_code)
            return Response({'name': employee.name, 'pay_code':pay_code}, status=status.HTTP_200_OK)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)







from rest_framework import viewsets
from .models import EmployeeLevelAssignment
from .serializers import EmployeeLevelAssignmentSerializer

class EmployeeLevelAssignmentViewSet(viewsets.ModelViewSet):
    queryset = EmployeeLevelAssignment.objects.all()
    serializer_class = EmployeeLevelAssignmentSerializer







from rest_framework import viewsets
from .models import LevelTwoTraineeInfo
from .serializers import NestedLevelTwoTraineeInfoSerializer

class NestedLevelTwoTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTraineeInfo.objects.all()
    serializer_class = NestedLevelTwoTraineeInfoSerializer






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelTwoTraineeInfo
from .serializers import NestedLevelTwoTraineeInfoSerializer

class GetTraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelTwoTraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelTwoTraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelTwoTraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)







from rest_framework import viewsets
from .models import LevelTwoQuality, LevelTwoQualityLine
from .serializers import LevelTwoQualitySerializer, LevelTwoQualityLineSerializer

class LevelTwoQualityViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQuality.objects.all()
    serializer_class = LevelTwoQualitySerializer


class LevelTwoQualityLineViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQualityLine.objects.all()
    serializer_class = LevelTwoQualityLineSerializer






from rest_framework import viewsets
from .models import (
    LevelTwoQATraineeInfo,
    LevelTwoQATrainingTopic,
    LevelTwoQAOJTDay,
    LevelTwoQAOJTScore,
)
from .serializers import (
    LevelTwoQATraineeInfoSerializer,
    LevelTwoQATrainingTopicSerializer,
    LevelTwoQAOJTDaySerializer,
    LevelTwoQAOJTScoreSerializer,
)


class LevelTwoQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATraineeInfo.objects.all()
    serializer_class = LevelTwoQATraineeInfoSerializer


class LevelTwoQATrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATrainingTopic.objects.all()
    serializer_class = LevelTwoQATrainingTopicSerializer


class LevelTwoQAOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQAOJTDay.objects.all()
    serializer_class = LevelTwoQAOJTDaySerializer


class LevelTwoQAOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQAOJTScore.objects.all()
    serializer_class = LevelTwoQAOJTScoreSerializer






from rest_framework import viewsets
from .models import LevelTwoQATraineeInfo
from .serializers import NestedLevelTwoQATraineeInfoSerializer

class NestedLevelTwoQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATraineeInfo.objects.all()
    serializer_class = NestedLevelTwoQATraineeInfoSerializer






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelTwoQATraineeInfo
from .serializers import NestedLevelTwoQATraineeInfoSerializer

# URL: /api/qa-trainee/<trainee_id>/<line_id>/

class GetQATraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelTwoQATraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelTwoQATraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelTwoQATraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    # def get(self, request, trainee_id, line_id):
    #     try:
    #         trainee = LevelTwoQATraineeInfo.objects.get(traineeId=trainee_id, line_id=line_id)
    #     except LevelTwoQATraineeInfo.DoesNotExist:
    #         return Response({"error": "Trainee not found with this ID and line"}, status=status.HTTP_404_NOT_FOUND)

    #     serializer = NestedLevelTwoQATraineeInfoSerializer(trainee)
    #     return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework import viewsets
from .models import LevelThreeProduction, LevelThreeLine, LevelThreeSubStation
from .serializers import (
    LevelThreeProductionSerializer,
    LevelThreeLineSerializer,
    LevelThreeSubStationSerializer
)


class LevelThreeProductionViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeProduction.objects.all()
    serializer_class = LevelThreeProductionSerializer


class LevelThreeLineViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeLine.objects.all()
    serializer_class = LevelThreeLineSerializer


class LevelThreeSubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeSubStation.objects.all()
    serializer_class = LevelThreeSubStationSerializer













from django.shortcuts import render

# Create your views here.

from rest_framework import viewsets
from .models import LevelThreeTraineeInfo, LevelThreeTrainingTopic, LevelThreeOJTDay, LevelThreeOJTScore
from .serializers import (
    LevelThreeTraineeInfoSerializer,
    LevelThreeTrainingTopicSerializer,
    LevelThreeOJTDaySerializer,
    LevelThreeOJTScoreSerializer,
)

class LevelThreeTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTraineeInfo.objects.all()
    serializer_class = LevelThreeTraineeInfoSerializer

class LevelThreeTrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTrainingTopic.objects.all()
    serializer_class = LevelThreeTrainingTopicSerializer

class LevelThreeOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeOJTDay.objects.all()
    serializer_class = LevelThreeOJTDaySerializer




from .utils import check_and_update_operator_skill_level_three

class LevelThreeOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeOJTScore.objects.all()
    serializer_class = LevelThreeOJTScoreSerializer

    def perform_create(self, serializer):
    # Save the new OJT score
     instance = serializer.save()

    # Update training status
     instance.trainee.calculate_and_save_training_status()

    # ✅ Call the Level 3 operator skill update
     if instance.trainee.trainee_Id:
         check_and_update_operator_skill_level_three(instance.trainee.trainee_Id)

    def get_queryset(self):
        trainee_name = self.request.query_params.get('trainee')
        if trainee_name:
            return LevelThreeOJTScore.objects.filter(trainee__trainee_name=trainee_name)
        return super().get_queryset()
    




from rest_framework import viewsets
from .models import LevelThreeTraineeInfo
from .serializers import NestedLevelThreeTraineeInfoSerializer

class NestedLevelThreeTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTraineeInfo.objects.all()
    serializer_class = NestedLevelThreeTraineeInfoSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelThreeTraineeInfo
from .serializers import NestedLevelThreeTraineeInfoSerializer

# URL: /api/levelthree-trainee/<trainee_id>/<station_id>/

class GetLevelThreeTraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelThreeTraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelThreeTraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelThreeTraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

    # def get(self, request, trainee_id, station_id):
    #     try:
    #         trainee = LevelThreeTraineeInfo.objects.get(trainee_Id=trainee_id, station_id=station_id)
    #     except LevelThreeTraineeInfo.DoesNotExist:
    #         return Response(
    #             {"error": "Trainee not found with this ID and station"},
    #             status=status.HTTP_404_NOT_FOUND
    #         )

    #     serializer = NestedLevelThreeTraineeInfoSerializer(trainee)
    #     return Response(serializer.data, status=status.HTTP_200_OK)















from rest_framework import viewsets
from .models import LevelThreeQuality, LevelThreeQualityLine
from .serializers import LevelThreeQualitySerializer, LevelThreeQualityLineSerializer

class LevelThreeQualityViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQuality.objects.all()
    serializer_class = LevelThreeQualitySerializer


class LevelThreeQualityLineViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQualityLine.objects.all()
    serializer_class = LevelThreeQualityLineSerializer







from rest_framework import viewsets
from .models import (
    LevelThreeQATraineeInfo,
    LevelThreeQATrainingTopic,
    LevelThreeQAOJTDay,
    LevelThreeQAOJTScore,
)
from .serializers import (
    LevelThreeQATraineeInfoSerializer,
    LevelThreeQATrainingTopicSerializer,
    LevelThreeQAOJTDaySerializer,
    LevelThreeQAOJTScoreSerializer,
)

class LevelThreeQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATraineeInfo.objects.all()
    serializer_class = LevelThreeQATraineeInfoSerializer


class LevelThreeQATrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATrainingTopic.objects.all()
    serializer_class = LevelThreeQATrainingTopicSerializer


class LevelThreeQAOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQAOJTDay.objects.all()
    serializer_class = LevelThreeQAOJTDaySerializer


class LevelThreeQAOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQAOJTScore.objects.all()
    serializer_class = LevelThreeQAOJTScoreSerializer








from rest_framework import viewsets
from .models import LevelThreeQATraineeInfo
from .serializers import NestedLevelThreeQATraineeInfoSerializer

class NestedLevelThreeQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATraineeInfo.objects.all()
    serializer_class = NestedLevelThreeQATraineeInfoSerializer






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelThreeQATraineeInfo
from .serializers import NestedLevelThreeQATraineeInfoSerializer

# URL: /api/qa-trainee/<trainee_id>/<line_id>/

class GetThreeQATraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelThreeQATraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelThreeQATraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelThreeQATraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)
#    def get(self, request, trainee_id, line_id):
#         try:
#             trainee = LevelThreeQATraineeInfo.objects.get(traineeId=trainee_id, line_id=line_id)
#         except LevelThreeQATraineeInfo.DoesNotExist:
#             return Response({"error": "Trainee not found with this ID and line"}, status=status.HTTP_404_NOT_FOUND)

#         serializer = NestedLevelThreeQATraineeInfoSerializer(trainee)
#         return Response(serializer.data, status=status.HTTP_200_OK)









from rest_framework import viewsets
from .models import ARVRTrainingContent
from .serializers import ARVRTrainingContentSerializer

class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ARVRTrainingContent.objects.all()
    serializer_class = ARVRTrainingContentSerializer






from rest_framework import viewsets
from .models import MCQQuestion
from .serializers import MCQQuestionSerializer

class MCQQuestionViewSet(viewsets.ModelViewSet):
    queryset = MCQQuestion.objects.all()
    serializer_class = MCQQuestionSerializer







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster

class EmployeeNameByCodeAPIView(APIView):
    def get(self, request, pay_code):
        try:
            employee = EmployeeMaster.objects.get(pay_code=pay_code)
            return Response({'name': employee.name}, status=status.HTTP_200_OK)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from tablib import Dataset
from .resources import BiometricAttendanceResource

class ExcelUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        file_format = 'xls' if file_obj.name.endswith('.xls') else 'xlsx'
        dataset = Dataset()
        try:
            imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        resource = BiometricAttendanceResource()
        result = resource.import_data(dataset, dry_run=True)

        if result.has_errors():
            errors = []
            for row_number, row_errors in result.row_errors():
                for error in row_errors:
                    errors.append(f"Row {row_number}: {str(error.error)}")
            return Response({'error': 'Import failed', 'details': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Perform actual import
        resource.import_data(dataset, dry_run=False)
        return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)






import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from tablib import Dataset
from .resources import BiometricAttendanceResource

class ExcelUploadFromPathView(APIView):
    # Set the path to your Excel file here
    EXCEL_FILE_PATH = r"E:\attendance.xlsx"  # Change this to your actual path

    def post(self, request, format=None):
        if not os.path.exists(self.EXCEL_FILE_PATH):
            return Response({'error': 'File not found on server path'}, status=status.HTTP_400_BAD_REQUEST)

        file_format = 'xls' if self.EXCEL_FILE_PATH.endswith('.xls') else 'xlsx'

        dataset = Dataset()
        try:
            with open(self.EXCEL_FILE_PATH, 'rb') as file_obj:
                imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        resource = BiometricAttendanceResource()
        result = resource.import_data(imported_data, dry_run=True)

        if result.has_errors():
            errors = []
            for row_number, row_errors in result.row_errors():
                for error in row_errors:
                    errors.append(f"Row {row_number}: {str(error.error)}")
            return Response({'error': 'Import failed', 'details': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Perform actual import
        resource.import_data(imported_data, dry_run=False)
        return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)




from rest_framework import viewsets
from .models import BiometricAttendance
from .serializers import BiometricAttendanceSerializer

class BiometricAttendanceViewSet(viewsets.ModelViewSet):
    queryset = BiometricAttendance.objects.all()
    serializer_class = BiometricAttendanceSerializer







from rest_framework import viewsets
from .models import MultiSkilling
from .serializers import NewMultiSkillingSerializer

class NewMultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = MultiSkilling.objects.all()
    serializer_class = NewMultiSkillingSerializer



from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithActiveSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only active skills for this employee
            active_skills = MultiSkilling.objects.filter(
                employee=emp, status='active'
            ).select_related('skill_level', 'station')

            skills = [
                {
                    "skill": skill.skill,
                    "skill_level": skill.skill_level.skill_level,
                    "start_date": skill.start_date,
                    "end_date": skill.end_date,
                    "notes": skill.notes,
                    "status": skill.status,
                }
                for skill in active_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)












from rest_framework.views import APIView
from rest_framework.response import Response
from .models import MultiSkilling, EmployeeMaster
from django.db.models import Prefetch, Q

class GroupedEmployeeSkillsView(APIView):
    def get(self, request):
        valid_statuses = ['scheduled', 'inprogress', 'completed']
        skills = MultiSkilling.objects.filter(status__in=valid_statuses) \
            .select_related('employee', 'skill_level', 'station') \
            .order_by('employee_id', 'start_date')

        grouped_data = {}

        for skill in skills:
            emp = skill.employee
            emp_id = emp.id

            if emp_id not in grouped_data:
                grouped_data[emp_id] = {
                    "employee_id": emp.id,
                    "pay_code": emp.pay_code,
                    "card_no": emp.card_no,
                    "name": emp.name,
                    "department": emp.department,
                    "section": emp.section,
                    "joining_date": emp.joining_date,
                    "skills": []
                }

            grouped_data[emp_id]["skills"].append({
                "station": skill.station.skill if skill.station else None,
                "station_number": skill.station.station_number if skill.station else None,
                "start_date": skill.start_date,
                "end_date": skill.end_date,
                "notes": skill.notes,
                "status": skill.status,
                "skill_level": skill.skill_level.name if skill.skill_level else None
            })

        return Response(list(grouped_data.values()))

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithCompletedSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only active skills for this employee
            active_skills = MultiSkilling.objects.filter(
                employee=emp, status='completed'
            ).select_related('skill_level', 'station')

            skills = [
                {
                    "skill": skill.skill,
                    "skill_level": skill.skill_level.skill_level,
                    "start_date": skill.start_date,
                    "end_date": skill.end_date,
                    "notes": skill.notes,
                    "status": skill.status,
                }
                for skill in active_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)














from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status as http_status
from .serializers import RefreshMultiSkillingSerializer
from .models import MultiSkilling

@api_view(['POST'])
def create_rescheduled_multiskilling(request):
    data = request.data.copy()  # Copy the request data
    data['status'] = 'rescheduled'  # Force status to 'rescheduled'

    serializer = RefreshMultiSkillingSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=http_status.HTTP_201_CREATED)
    return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)






from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithRescheduledSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only 'rescheduled' skills for this employee
            rescheduled_skills = MultiSkilling.objects.filter(
                employee=emp, status='rescheduled'
            ).select_related('skill_level', 'station')

            # If the employee has no rescheduled skills, skip
            if not rescheduled_skills.exists():
                continue

            skills = [
                {
                    "id": skill.id,
                    "skill": skill.skill,
                    "notes": skill.notes,
                    "status": skill.status,
                    "reason": skill.reason,
                    "refreshment_date": skill.refreshment_date,
                }
                for skill in rescheduled_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)









from rest_framework import viewsets
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class TrainingReportViewSet(viewsets.ModelViewSet):
    queryset = TrainingReport.objects.all().order_by('-month')
    serializer_class = TrainingReportSerializer



from rest_framework import viewsets
from .models import UnifiedDefectReport
from .serializers import UnifiedDefectReportSerializer

class UnifiedDefectReportViewSet(viewsets.ModelViewSet):
    queryset = UnifiedDefectReport.objects.all().order_by('-month')
    serializer_class = UnifiedDefectReportSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TrainingReport
from .serializers import TrainingReportSerializer
from django.db.models import Sum

class TrainingSummaryView(APIView):
    def get(self, request):
        summary = TrainingReport.objects.aggregate(
            new_operators_joined=Sum("new_operators_joined"),
            new_operators_trained=Sum("new_operators_trained"),
            total_trainings_planned=Sum("total_trainings_planned"),
            total_trainings_actual=Sum("total_trainings_actual")
        )
        return Response(summary)





from rest_framework.generics import ListAPIView
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class OperatorsJoinedVsTrainedView(ListAPIView):
    queryset = TrainingReport.objects.all().order_by("month")
    serializer_class = TrainingReportSerializer




from rest_framework.views import APIView
from .models import UnifiedDefectReport
from rest_framework.response import Response

class MSILDefectsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='MSIL').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])





class CTQDefectsAllPlantsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='All Plants').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])




class InternalRejectionView(APIView):
    def get(self, request):
        internal = UnifiedDefectReport.objects.all()
        rejection = internal.aggregate(
            total_internal_rejection=Sum('total_internal_rejection'),
            ctq_internal_rejection=Sum('ctq_internal_rejection')
        )
        return Response(rejection)



from rest_framework import viewsets
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class TrainingReportViewSet(viewsets.ModelViewSet):
    queryset = TrainingReport.objects.all().order_by('-month')
    serializer_class = TrainingReportSerializer



from rest_framework import viewsets
from .models import UnifiedDefectReport
from .serializers import UnifiedDefectReportSerializer

class UnifiedDefectReportViewSet(viewsets.ModelViewSet):
    queryset = UnifiedDefectReport.objects.all().order_by('-month')
    serializer_class = UnifiedDefectReportSerializer

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer


class SubtopicWiseTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

    @action(detail=False, methods=['get'], url_path='(?P<id>\d+)')
    def subtopicwise(self, request, id=None):
        contents = TrainingContent.objects.filter(subtopic_content_id=id)
        serializer = self.get_serializer(contents, many=True)
        return Response(serializer.data)
    


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MCQQuestion
from .serializers import MCQQuestionSerializer

class MCQBySubtopicView(APIView):
    def get(self, request, subtopic_id):
        mcqs = MCQQuestion.objects.filter(subtopic_content_id=subtopic_id)
        serializer = MCQQuestionSerializer(mcqs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

# import xlrd
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster
from .serializers import ExcelUploadSerializer
import xlrd


def parse_excel_date(value, datemode):
    if isinstance(value, float):  # Excel serial date
        return xlrd.xldate.xldate_as_datetime(value, datemode).date()
    elif isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d/%m/%Y").date()
        except:
            return None
    return None

class EmployeeExcelUploadView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            try:
                workbook = xlrd.open_workbook(file_contents=excel_file.read())
                sheet = workbook.sheet_by_index(0)

                # Normalize headers (lowercase, no leading/trailing whitespace), skip first column
                raw_headers = [str(cell.value).strip().lower() for cell in sheet.row(5)][1:]

                for row_idx in range(6, sheet.nrows):  # Start after header row
                    row_cells = sheet.row(row_idx)[1:]  # Skip "Sr. No." column
                    row_data = {raw_headers[i]: row_cells[i].value for i in range(len(raw_headers))}

                    # Safely fetch and convert dates
                    birth_date = parse_excel_date(row_data.get('birth date'), workbook.datemode)
                    joining_date = parse_excel_date(row_data.get('joining date'), workbook.datemode)

                    if not birth_date or not joining_date:
                        continue  # Skip rows with invalid dates

                    # Create or update the employee record
                    EmployeeMaster.objects.update_or_create(
                        pay_code=row_data.get('pay code'),
                        defaults={
                            'card_no': row_data.get('card no.'),
                            'sex': row_data.get('sex'),
                            'birth_date': birth_date,
                            'name': row_data.get('name'),
                            'guardian_name': row_data.get("guardian's name", ''),
                            'department': row_data.get('department'),
                            'section': row_data.get('section'),
                            'desig_category': row_data.get('desig/categor') or None,
                            'joining_date': joining_date,
                            'auth_shift': row_data.get('auth shift'),
                            'shift_type': row_data.get('shift type'),
                            'shift_pattern': row_data.get('shift pattern'),
                            'first_weekly_off': row_data.get('1st weekly') or '',  # corrected key
                            'second_weekly_off': None,
                            'second_weekly_off_fh': None,
                            'ot_allowed_rate': False,
                            'round_the_clock': False,
                        }
                    )

                return Response({"message": "Employees uploaded successfully"}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)






# from rest_framework import viewsets
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from .models import MachineAllocation, OperatorSkill, Machine
# from .serializers import EmployeeNameSerializer

# class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
#     queryset = MachineAllocation.objects.all()
#     serializer_class = ...  # Your main serializer

#     @action(detail=False, methods=['get'], url_path='eligible-employees')
#     def eligible_employees(self, request):
#         machine_id = request.query_params.get('machine_id')
#         if not machine_id:
#             return Response({'error': 'machine_id is required'}, status=400)

#         try:
#             machine = Machine.objects.get(id=machine_id)
#         except Machine.DoesNotExist:
#             return Response({'error': 'Machine not found'}, status=404)

#         matching_skills = OperatorSkill.objects.filter(station__skill=machine.process)
#         employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
#         employees = EmployeeMaster.objects.filter(id__in=employee_ids)

#         serializer = EmployeeNameSerializer(employees, many=True)
#         return Response(serializer.data)


# easytest


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from rest_framework import viewsets


class KeyEventCreateView(APIView):
    def post(self, request):
        serializer = KeyEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Key event saved'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LatestKeyEventView(APIView):
    def get(self, request):
        try:
            latest_event = KeyEvent.objects.latest('timestamp')
            serializer = KeyEventSerializer(latest_event)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyEvent.DoesNotExist:
            return Response({"message": "No key events yet."}, status=status.HTTP_404_NOT_FOUND)

        
# api/views.py
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ConnectEventSerializer

@api_view(['POST'])
def connect_event_create(request):
    serializer = ConnectEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




@api_view(['POST'])
def vote_event_create(request):
    serializer = VoteEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import Question, EmployeeMaster, Level, Station, TestSession, Score
from .serializers import QuestionSerializer, EmployeeSerializer, ScoreSerializer, SimpleScoreSerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Question
from .serializers import QuestionSerializer


from rest_framework import generics

class QuestionListCreateView(generics.ListCreateAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer

    def get_queryset(self):
        paper_id = self.request.query_params.get('paper_id')
        if paper_id:
            return self.queryset.filter(question_paper__id=paper_id)
        return self.queryset
    
from rest_framework import generics

class QuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer




class QuestionPaperListCreateView(generics.ListCreateAPIView):
    queryset = QuestionPaper.objects.all()
    serializer_class = QuestionPaperSerializer

class QuestionsByPaperView(generics.ListAPIView):
    serializer_class = QuestionSerializer

    def get_queryset(self):
        paper_id = self.kwargs.get('paper_id')
        return Question.objects.filter(question_paper_id=paper_id)




class EmployeeListCreateView(APIView):
    def get(self, request):
        employees = EmployeeMaster.objects.all()
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(session_key=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)


class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({s.key_id: s.employee.name for s in mapping})

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import EmployeeMaster, Station, TestSession, QuestionPaper

# class StartTestSessionView(APIView):
#     def post(self, request):
#         try:
#             test_name = request.data.get("test_name")
#             assignments = request.data.get("assignments", [])
#             question_paper_id = request.data.get("question_paper_id")

#             if not test_name or not assignments:
#                 return Response(
#                     {"error": "Test name and assignments are required."},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )

#             question_paper = None
#             if question_paper_id:
#                 question_paper = get_object_or_404(QuestionPaper, id=question_paper_id)

#             for item in assignments:
#                 key_id = item.get("key_id")
#                 employee_id = item.get("employee_id")

#                 if not key_id or not employee_id:
#                     return Response(
#                         {"error": "key_id and employee_id are required in each assignment."},
#                         status=status.HTTP_400_BAD_REQUEST,
#                     )

#                 employee = get_object_or_404(EmployeeMaster, id=employee_id)

#                 # You can skip station logic if not using skill anymore
#                 TestSession.objects.create(
#                     test_name=test_name,
#                     key_id=key_id,
#                     employee=employee,
#                     level=None,
#                     skill=None,  # Assuming nullable=True on model
#                     question_paper=question_paper,
#                 )

#             return Response({"status": "ok"}, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class StartTestSessionView(APIView):
    def post(self, request):
        try:
            print("Incoming Request Data:", request.data)

            test_name = request.data.get("test_name")
            assignments = request.data.get("assignments", [])
            question_paper_id = request.data.get("question_paper_id")
            level = request.data.get("level")  # string
            skill_id = request.data.get("skill")  # foreign key

            if not test_name or not assignments:
                response_data = {"error": "Test name and assignments are required."}
                print("Response:", response_data)
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            question_paper = None
            if question_paper_id:
                question_paper = get_object_or_404(QuestionPaper, id=question_paper_id)

            skill = None
            if skill_id:
                skill = get_object_or_404(Station, id=skill_id)

            for item in assignments:
                key_id = item.get("key_id")
                employee_id = item.get("employee_id")

                if not key_id or not employee_id:
                    response_data = {"error": "key_id and employee_id are required in each assignment."}
                    print("Response:", response_data)
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

                employee = get_object_or_404(EmployeeMaster, id=employee_id)

                TestSession.objects.create(
                    test_name=test_name,
                    key_id=key_id,
                    employee=employee,
                    level=level,  # now just saving the string directly
                    skill=skill,
                    question_paper=question_paper,
                )

            response_data = {"status": "ok"}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {"error": str(e)}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



##################################################################################################################################################


# from rest_framework.views import APIView
# from rest_framework.response import Response
# from .models import TestSession, Question, Score

# class EndTestSessionView(APIView):
#     def post(self, request):
#         try:
#             key_id_to_answers = request.data  # { key_id: [answers] }
#             results = []
#             test_name = ''

#             sessions = TestSession.objects.select_related('employee', 'skill', 'question_paper').all()

#             for session in sessions:
#                 key_id = str(session.key_id)
#                 employee = session.employee
#                 test_name = session.test_name
#                 question_paper = session.question_paper

#                 questions = list(Question.objects.filter(question_paper=question_paper).order_by('id'))
#                 answers = key_id_to_answers.get(key_id, [])

#                 correct_count = 0
#                 for i, ans in enumerate(answers):
#                     if i < len(questions) and ans == questions[i].correct_index:
#                         correct_count += 1

#                 percentage = round((correct_count / len(questions)) * 100) if questions else 0
#                 passed = percentage >= 80

#                 Score.objects.create(
#                     employee=employee,
#                     marks=correct_count,
#                     percentage=percentage,
#                     passed=passed,
#                     test_name=test_name,
#                     level=None,
#                     skill=session.skill
#                 )

#                 results.append({
#                     'name': employee.name,
#                     'marks': correct_count,
#                     'percentage': percentage,
#                     'passed': passed
#                 })

#             TestSession.objects.all().delete()

#             return Response({'test_name': test_name, 'results': results}, status=200)

#         except Exception as e:
#             return Response({'error': str(e)}, status=500)
####################################################################################################################
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, LevelTwoTraineeInfo
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


class EndTestSessionView(APIView):
    def post(self, request):
        try:
            key_id_to_answers = request.data  # { key_id: [answers] }
            results = []

            sessions = TestSession.objects.select_related(
                'employee', 'skill', 'question_paper'
            ).all()

            for session in sessions:
                key_id = str(session.key_id)
                if key_id not in key_id_to_answers:
                    continue

                employee = session.employee
                test_name = session.test_name
                question_paper = session.question_paper
                level = session.level
                skill = session.skill

                questions = list(
                    Question.objects.filter(question_paper=question_paper).order_by('id')
                )
                answers = key_id_to_answers.get(key_id, [])

                correct_count = 0
                for i, ans in enumerate(answers):
                    if i < len(questions) and ans == questions[i].correct_index:
                        correct_count += 1

                percentage = round((correct_count / len(questions)) * 100) if questions else 0
                passed = percentage >= 80

                # ✅ Create the Score
                Score.objects.create(
                    employee=employee,
                    marks=correct_count,
                    percentage=percentage,
                    passed=passed,
                    test_name=test_name,
                    test=session,
                    level=level,
                    skill=skill
                )

                # ✅ Call skill updater only if passed
                if passed:
                    try:
                        trainee = LevelTwoTraineeInfo.objects.get(traineeId=employee.pay_code)
                        trainee.calculate_and_save_training_status()

                        # Call both functions
                        check_and_update_operator_skill(trainee.traineeId)
                        check_and_update_operator_skill_level_three(trainee.traineeId)

                    except LevelTwoTraineeInfo.DoesNotExist:
                        pass  # Skip if trainee not found

                results.append({
                    'name': employee.name,
                    'marks': correct_count,
                    'percentage': percentage,
                    'passed': passed
                })

            # ✅ Delete sessions
            TestSession.objects.all().delete()

            return Response({
                'test_name': test_name,
                'results': results
            }, status=200)

        except Exception as e:
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)










class PastTestSessionsView(APIView):
    def get(self, request):
        qs = Score.objects.values('test_name').distinct()
        return Response([s['test_name'] for s in qs])


class ScoresByTestView(APIView):
    def get(self, request, name):
        scores = (
            Score.objects
            .filter(test_name=name)
            .select_related('employee', 'skill')  # ✅ FIX: Removed 'level'
        )

        questions_count = Question.objects.count() or 1

        data = []
        for s in scores:
            data.append({
                'employee_id': s.employee.id,
                'name': s.employee.name,
                'marks': s.marks,
                'percentage': s.percentage,
                'level_name': s.level if s.level else '',  # ✅ FIX: removed .name
                'skill': s.skill.skill if s.skill else '',  # assuming Station.skill is a string
                'section': s.employee.section if s.employee.section else '',
            })
        print("GET /api/scores-by-session/ response:", data) 
        return Response(data)

    
class ResultSummaryAPIView(APIView):
    def get(self, request):
        scores = Score.objects.select_related('employee', 'level', 'skill')
        data = []
        for score in scores:
            percentage = round((score.marks / 10) * 100, 2)  # Adjust total marks accordingly
            result = 'Pass' if score.marks >= 8 else 'Retraining' if score.marks >= 5 else 'Fail'

            data.append({
                "employee_id": score.employee.id,
                "name": score.employee.name,
                "marks": score.marks,
                "percentage": percentage,
                "section": score.employee.section,  # assuming CharField
                "level_name": score.level.name if score.level else '',
                "skill": score.skill.skill if score.skill else '',  # Station.skill string
                "result": result,
            })

        serializer = SimpleScoreSerializer(data, many=True)
        return Response(serializer.data)


class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('skill', flat=True).distinct()
        return Response(skills)


class ScoresBySessionView(APIView):
    def get(self, request, session_key):
        scores = Score.objects.filter(session_key=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





1

#Employee Card 



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import OperatorSkill, EmployeeMaster
from .serializers import OperatorCardSkillSerializer

class OperatorSkillByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if name:
            try:
                employee = EmployeeMaster.objects.get(name=name)
                operator_skills = OperatorSkill.objects.filter(operator=employee)
                serializer = OperatorCardSkillSerializer(operator_skills, many=True)
                return Response(serializer.data)
            except EmployeeMaster.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, Score
from .serializers import CardScoreSerializer

class ScoreByEmployeeNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if name:
            try:
                employee = EmployeeMaster.objects.get(name=name)
                scores = Score.objects.filter(test__employee=employee)
                serializer = CardScoreSerializer(scores, many=True)
                return Response(serializer.data)
            except EmployeeMaster.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, MultiSkilling
from .serializers import CardMultiSkillingSerializer

class MultiSkillingByEmployeeView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            employee = EmployeeMaster.objects.get(name=name)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        records = MultiSkilling.objects.filter(employee=employee)
        serializer = CardMultiSkillingSerializer(records, many=True)
        return Response(serializer.data)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, RefreshmentTraining
from .serializers import CardRefreshmentTrainingSerializer

class RefreshmentTrainingByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            employee = EmployeeMaster.objects.get(name=name)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        records = RefreshmentTraining.objects.filter(employee=employee)
        serializer = CardRefreshmentTrainingSerializer(records, many=True)
        return Response(serializer.data)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster
from .serializers import CardEmployeeMasterSerializer

class CardEmployeeDetailByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = EmployeeMaster.objects.get(name=name)
            serializer = CardEmployeeMasterSerializer(employee)
            return Response(serializer.data)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)










from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView

from .models import (
    EmployeeMaster,
    OperatorSkill,
    Score,
    MultiSkilling,
    RefreshmentTraining
)

from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    CardMultiSkillingSerializer,
    CardRefreshmentTrainingSerializer
)

class EmployeeCardDetailsView(APIView):
    def get(self, request):
        card_no = request.query_params.get('card_no')
        if not card_no:
            return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = EmployeeMaster.objects.get(card_no=card_no)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        # Fetch and serialize all related data
        employee_data = CardEmployeeMasterSerializer(employee).data
        operator_skills = OperatorCardSkillSerializer(OperatorSkill.objects.filter(operator=employee), many=True).data
        scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
        multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data
        refreshment_training = CardRefreshmentTrainingSerializer(RefreshmentTraining.objects.filter(employee=employee), many=True).data

        # Construct full response
        response_data = {
            'employee': employee_data,
            'operator_skills': operator_skills,
            'scores': scores,
            'multi_skilling': multi_skilling,
            'refreshment_training': refreshment_training,
        }

        # Print to console
        print("==== Employee Card Details ====")
        import pprint
        pprint.pprint(response_data)  # pretty-print for readability
        print("================================")

        return Response(response_data)



















from .models import HanContent, HanTrainingContent
from .serializers import HanContentSerializer, HanTrainingContentSerializer


class HanContentViewSet(viewsets.ModelViewSet):
    queryset = HanContent.objects.all()
    serializer_class = HanContentSerializer


class HanTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = HanTrainingContent.objects.all()
    serializer_class = HanTrainingContentSerializer


class HanTrainingContentCreateView(APIView):
    def post(self, request):
        serializer = HanTrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HanTrainingContent
from .serializers import HanTrainingContentSerializer

class HanTrainingContentByContentID(APIView):
    def get(self, request, han_content_id):
        training_contents = HanTrainingContent.objects.filter(han_content_id=han_content_id)
        serializer = HanTrainingContentSerializer(training_contents, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from .models import ShoContent, ShoTrainingContent
from .serializers import ShoContentSerializer, ShoTrainingContentSerializer


class ShoContentViewSet(viewsets.ModelViewSet):
    queryset = ShoContent.objects.all()
    serializer_class = ShoContentSerializer


class ShoTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ShoTrainingContent.objects.all()
    serializer_class = ShoTrainingContentSerializer


class ShoTrainingContentCreateView(APIView):
    def post(self, request):
        serializer = ShoTrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HanTrainingContent
from .serializers import HanTrainingContentSerializer

class ShoTrainingContentByContentID(APIView):
    def get(self, request, sho_content_id):  # Make sure this matches the URL parameter name
        training_contents = ShoTrainingContent.objects.filter(sho_content_id=sho_content_id)
        serializer = ShoTrainingContentSerializer(training_contents, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



 
# employeereportpdfview
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import json
import traceback
from .models import EmployeeMaster, OperatorSkill, Score, MultiSkilling, RefreshmentTraining

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            print(f"Processing card_no: {card_no}")

            # 2. Get employee record
            try:
                employee = EmployeeMaster.objects.get(card_no=card_no)
                print(f"Found employee: {employee.name}")
            except EmployeeMaster.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # 3. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = self.create_pdf_content(employee)
            
            # 4. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_card_number(self, request):
        """Helper method to extract card_no from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')

    def create_pdf_content(self, employee):
        """Generate the PDF content structure"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        story.append(Paragraph(f"Employee Comprehensive Report", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_basic_info(story, styles, employee)
        self._add_operator_skills(story, styles, employee)
        self._add_scores(story, styles, employee)
        self._add_multi_skills(story, styles, employee)
        self._add_refreshment_training(story, styles, employee)
        
        return story


    def _get_table_style(self):
        """Returns a consistent, professional style for all tables"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),  # Steel blue
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 4),
            
            # Data row styling
            ('ALIGN', (0,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            
            # Cell padding
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])

    def _add_basic_info(self, story, styles, employee):
        """Employee basic information with improved layout"""
        story.append(Paragraph("Basic Information", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        basic_data = [
            ["Field", "Value"],
            ["Name:", employee.name],
            ["Card No:", employee.card_no],
            ["Department:", employee.department],
            ["Section:", employee.section],
            ["Designation:", employee.desig_category],
            ["Joining Date:", employee.joining_date.strftime('%Y-%m-%d') if employee.joining_date else "N/A"],
            ["Gender:", employee.sex],
            ["Birth Date:", employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else "N/A"],
            ["Guardian:", employee.guardian_name or "N/A"]
        ]
        
        basic_table = Table(basic_data, colWidths=[150, 300])
        style = TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (1,0), 'CENTER'),
            ('FONTNAME', (0,0), (1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (1,0), 10),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ])
        basic_table.setStyle(style)
        story.append(basic_table)
        story.append(Spacer(1, 24))

    def _add_operator_skills(self, story, styles, employee):
        """Operator skills with professional table design"""
        skills = OperatorSkill.objects.filter(operator=employee).select_related('station')
        if not skills.exists():
            return
            
        story.append(Paragraph("Operator Skills", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        skill_data = [["Station", "Skill Level"]]
        for skill in skills:
            skill_data.append([
                str(skill.station),
                skill.skill_level or "N/A"
            ])
        
        skills_table = Table(skill_data, colWidths=[250, 100])
        skills_table.setStyle(self._get_table_style())
        story.append(skills_table)
        story.append(Spacer(1, 24))


    def _add_scores(self, story, styles, employee):
        """Scores table with proper result formatting"""
        scores = Score.objects.filter(employee=employee)
        if not scores.exists():
            return
            
        story.append(Paragraph("Scores and Assessments", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        score_data = [["Test", "Marks", "%", "Result", "Date"]]
        for score in scores:
            score_data.append([
                score.test_name or "N/A",
                str(score.marks) if score.marks is not None else "N/A",
                f"{score.percentage}%" if score.percentage is not None else "N/A",
                "Pass" if score.passed else "Fail",
                score.created_at.strftime("%d-%b-%Y") if score.created_at else "N/A"
            ])
        
        scores_table = Table(score_data, colWidths=[150, 60, 60, 80, 80])
        style = self._get_table_style()
        style.add('ALIGN', (1,1), (2,-1), 'RIGHT')
        
        # Add result coloring
        for row in range(1, len(score_data)):
            result = scores[row-1].passed  
            for style_command in self._get_result_badge_style(result, row):
                style.add(*style_command)
        
        scores_table.setStyle(style)
        story.append(scores_table)
        story.append(Spacer(1, 24))

    def _add_multi_skills(self, story, styles, employee):
        """Multi-skilling with proper status formatting"""
        multi_skills = MultiSkilling.objects.filter(employee=employee).select_related('station', 'skill_level')
        if not multi_skills.exists():
            return
            
        story.append(Paragraph("Multi-Skilling", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        multi_data = [["Skill", "Status", "Station", "Level", "Start", "End"]]
        for skill in multi_skills:
            multi_data.append([
                skill.skill or "N/A",
                skill.status.capitalize() if skill.status else "N/A",
                str(skill.station) if skill.station else "N/A",
                skill.skill_level.skill_level if skill.skill_level else "N/A",
                skill.start_date.strftime('%d-%b-%Y') if skill.start_date else "N/A",
                skill.end_date.strftime('%d-%b-%Y') if skill.end_date else "N/A"
            ])
        
        multi_table = Table(multi_data, colWidths=[120, 80, 100, 60, 80, 80])
        style = self._get_table_style()
        
        # Add status coloring
        for row in range(1, len(multi_data)):
            status = multi_skills[row-1].status
            for style_command in self._get_status_badge_style(status, row):
                style.add(*style_command)
        
        multi_table.setStyle(style)
        story.append(multi_table)
        story.append(Spacer(1, 24))

    def _get_result_badge_style(self, result, row):
        """Returns properly structured style commands for result badges"""
        if isinstance(result, bool):
            color = colors.green if result else colors.red
        else:
            color = colors.green if str(result).lower() == "pass" else colors.red
        
        # Return a complete style command tuple
        return [
            ('TEXTCOLOR', (3, row), (3, row), color),
            ('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        ]

    def _get_status_badge_style(self, status, row):
        """Returns properly structured style commands for status badges"""
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.gray,
            'rescheduled': colors.darkblue, 
        }
        color = status_colors.get(status.lower(), colors.black)
        return [
            ('TEXTCOLOR', (1, row), (1, row), color),
            ('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
        ]

    def _add_refreshment_training(self, story, styles, employee):
        """Refreshment training with professional layout"""
        trainings = RefreshmentTraining.objects.filter(employee=employee).select_related(
            'station', 'skill', 'skill_level'
        )
        if not trainings.exists():
            return
            
        story.append(Paragraph("Refreshment Training", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        training_data = [["Skill", "Station", "Level", "Start Date", "End Date", "Reason"]]
        for training in trainings:
            training_data.append([
                training.skill.skill if training.skill else "N/A",
                str(training.station) if training.station else "N/A",
                training.skill_level.skill_level if training.skill_level else "N/A",
                training.start_date.strftime('%d-%b-%Y') if training.start_date else "N/A",
                training.end_date.strftime('%d-%b-%Y') if training.end_date else "N/A",
                training.reason_for_refreshment or "N/A"
            ])
        
        training_table = Table(training_data, colWidths=[120, 100, 60, 80, 80, 150])
        style = self._get_table_style()
        style.add('ALIGN', (5,1), (5,-1), 'LEFT') 
        training_table.setStyle(style)
        story.append(training_table)



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MainDepartment
from .serializers import MainDepartmentSerializer

class MainDepartmentListView(APIView):
    def get(self, request):
        departments = MainDepartment.objects.all()
        serializer = MainDepartmentSerializer(departments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MainLine
from .serializers import MainLineByDepartmentSerializer

class MainLinesByDepartmentView(APIView):
    def get(self, request, department_id):
        main_lines = MainLine.objects.filter(department_id=department_id)
        serializer = MainLineByDepartmentSerializer(main_lines, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SubLine
from .serializers import SubLineByMainLineSerializer,StationByLineSerializer

class SubLinesByMainLineView(APIView):
    def get(self, request, main_line_id):
        sub_lines = SubLine.objects.filter(main_line_id=main_line_id)
        serializer = SubLineByMainLineSerializer(sub_lines, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)







class StationsBySubLineView(APIView):
    def get(self, request, sub_line_id):
        stations = Station.objects.filter(sub_line_id=sub_line_id)
        serializer = StationByLineSerializer(stations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import json
import traceback
from .models import EmployeeMaster, OperatorSkill, Score, MultiSkilling, RefreshmentTraining

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            print(f"Processing card_no: {card_no}")

            # 2. Get employee record
            try:
                employee = EmployeeMaster.objects.get(card_no=card_no)
                print(f"Found employee: {employee.name}")
            except EmployeeMaster.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # 3. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=50, bottomMargin=50, leftMargin=50, rightMargin=50)
            story = self.create_pdf_content(employee)
            
            # 4. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_card_number(self, request):
        """Helper method to extract card_no from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')

    def create_pdf_content(self, employee):
        """Generate the PDF content structure"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title with proper alignment
        title = Paragraph(f"Employee History Report", styles['Title'])
        title.alignment = 1  # Center alignment
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Add all sections
        self._add_basic_info(story, styles, employee)
        self._add_operator_skills(story, styles, employee)
        self._add_scores(story, styles, employee)
        self._add_multi_skills(story, styles, employee)
        self._add_refreshment_training(story, styles, employee)
        
        return story

    def _get_base_table_style(self):
        """Returns the base table style that all tables will use"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Data row styling
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Default all data to left alignment
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
            
            # Consistent padding for all cells
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])

    def _add_section_heading(self, story, styles, heading_text):
        """Add a consistent section heading"""
        heading = Paragraph(heading_text, styles['Heading2'])
        heading.alignment = 0  # Left alignment
        story.append(heading)
        story.append(Spacer(1, 10))

    def _add_basic_info(self, story, styles, employee):
        """Employee basic information with consistent layout"""
        self._add_section_heading(story, styles, "Basic Information")
        
        # Prepare data
        basic_data = [
            ["Field", "Value"],
            ["Name", employee.name or "N/A"],
            ["Card No", employee.card_no or "N/A"],
            ["Department", employee.department or "N/A"],
            ["Section", employee.section or "N/A"],
            ["Designation", employee.desig_category or "N/A"],
            ["Joining Date", employee.joining_date.strftime('%d-%b-%Y') if employee.joining_date else "N/A"],
            ["Gender", employee.sex or "N/A"],
            ["Birth Date", employee.birth_date.strftime('%d-%b-%Y') if employee.birth_date else "N/A"],
            ["Guardian", employee.guardian_name or "N/A"]
        ]
        
        # Create table with consistent width (500 points total)
        basic_table = Table(basic_data, colWidths=[150, 350])
        
        # Apply base style
        style = self._get_base_table_style()
        # Make field column bold
        style.add('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold')
        
        basic_table.setStyle(style)
        story.append(basic_table)
        story.append(Spacer(1, 20))

    def _add_operator_skills(self, story, styles, employee):
        """Operator skills with consistent layout"""
        skills = OperatorSkill.objects.filter(operator=employee).select_related('station')
        if not skills.exists():
            return
            
        self._add_section_heading(story, styles, "Operator Skills")
        
        # Prepare data
        skill_data = [["Station", "Skill Level"]]
        for skill in skills:
            skill_data.append([
                str(skill.station) if skill.station else "N/A",
                skill.skill_level or "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        skills_table = Table(skill_data, colWidths=[300, 200])
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align skill level column
        style.add('ALIGN', (1, 1), (1, -1), 'CENTER')
        
        skills_table.setStyle(style)
        story.append(skills_table)
        story.append(Spacer(1, 20))

    def _add_scores(self, story, styles, employee):
        """Scores table with consistent layout"""
        scores = Score.objects.filter(employee=employee)
        if not scores.exists():
            return
            
        self._add_section_heading(story, styles, "Scores and Assessments")
        
        # Prepare data
        score_data = [["Test Name", "Marks", "Percentage", "Result", "Date"]]
        for score in scores:
            score_data.append([
                score.test_name or "N/A",
                str(score.marks) if score.marks is not None else "N/A",
                f"{score.percentage}%" if score.percentage is not None else "N/A",
                "Pass" if score.passed else "Fail",
                score.created_at.strftime("%d-%b-%Y") if score.created_at else "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        scores_table = Table(score_data, colWidths=[150, 70, 90, 70, 120])
        
        # Apply base style
        style = self._get_base_table_style()
        # Right align numerical columns
        style.add('ALIGN', (1, 1), (1, -1), 'RIGHT')   # Marks
        style.add('ALIGN', (2, 1), (2, -1), 'RIGHT')   # Percentage
        # Center align result and date
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Result
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # Date
        
        # Add result coloring
        for row in range(1, len(score_data)):
            if row - 1 < len(scores):
                result = scores[row - 1].passed
                color = colors.green if result else colors.red
                style.add('TEXTCOLOR', (3, row), (3, row), color)
                style.add('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        
        scores_table.setStyle(style)
        story.append(scores_table)
        story.append(Spacer(1, 20))


    def _add_multi_skills(self, story, styles, employee):
        """Multi-skilling with consistent layout"""
        multi_skills = MultiSkilling.objects.filter(employee=employee).select_related('station', 'skill_level')
        if not multi_skills.exists():
            return
            
        self._add_section_heading(story, styles, "Multi-Skilling")
        
        # Prepare data with proper column distribution
        multi_data = [["Skill", "Status", "Station", "Level", "Start Date", "End Date"]]
        for skill in multi_skills:
            # Ensure skill name is clean and status is separate
            skill_name = (skill.skill or "N/A").strip()
            status = (skill.status or "N/A").strip().capitalize()
            
            multi_data.append([
                skill_name,
                status,
                str(skill.station) if skill.station else "N/A",
                skill.skill_level.skill_level if skill.skill_level else "N/A",
                skill.start_date.strftime('%d-%b-%Y') if skill.start_date else "N/A",
                skill.end_date.strftime('%d-%b-%Y') if skill.end_date else "N/A"
            ])
        
        # Create table with adjusted column widths
        # Total width should be around 500 (letter width minus margins)
        col_widths = [120, 80, 80, 50, 90, 80]  # Total = 500
        
        multi_table = Table(multi_data, colWidths=col_widths)
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align status, level, and dates
        style.add('ALIGN', (1, 1), (1, -1), 'CENTER')  # Status
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Level
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # Start Date
        style.add('ALIGN', (5, 1), (5, -1), 'CENTER')  # End Date
        
        # Add word wrapping for all columns
        style.add('WORDWRAP', (0, 0), (-1, -1), True)
        
        # Add status coloring
        for row in range(1, len(multi_data)):
            if row - 1 < len(multi_skills):
                status = multi_skills[row - 1].status
                if status:
                    color = self._get_status_color(status.lower())
                    style.add('TEXTCOLOR', (1, row), (1, row), color)
                    style.add('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
        
        multi_table.setStyle(style)
        story.append(multi_table)
        story.append(Spacer(1, 20))

    def _add_refreshment_training(self, story, styles, employee):
        """Refreshment training with consistent layout"""
        trainings = RefreshmentTraining.objects.filter(employee=employee).select_related(
            'station', 'skill', 'skill_level'
        )
        if not trainings.exists():
            return
            
        self._add_section_heading(story, styles, "Refreshment Training")
        
        # Prepare data
        training_data = [["Skill", "Station", "Level", "Start Date", "End Date", "Reason"]]
        for training in trainings:
            training_data.append([
                training.skill.skill if training.skill else "N/A",
                str(training.station) if training.station else "N/A",
                training.skill_level.skill_level if training.skill_level else "N/A",
                training.start_date.strftime('%d-%b-%Y') if training.start_date else "N/A",
                training.end_date.strftime('%d-%b-%Y') if training.end_date else "N/A",
                training.reason_for_refreshment or "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        training_table = Table(training_data, colWidths=[90, 90, 60, 80, 80, 100])
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align level and dates
        style.add('ALIGN', (2, 1), (2, -1), 'CENTER')  # Level
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Start Date
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # End Date
        
        training_table.setStyle(style)
        story.append(training_table)

    def _get_status_color(self, status):
        """Get appropriate color for status"""
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'in progress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.red,
            'rescheduled': colors.navy,
        }
        return status_colors.get(status.lower(), colors.black)
    

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import EmployeeMaster
from .serializers import EmployeeSerializer
from datetime import datetime

class EmployeeExcelViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMaster.objects.all()
    serializer_class = EmployeeSerializer
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "EMPLOYEE MASTER"

        ws.merge_cells('A1:S1')
        company_cell = ws['A1']
        company_cell.value = "Company Name: KRISHNA MARUTI SEAT -JOSHI SAI ENTERPRISES, NEHA ENTERPRISES, MADHU ENTERPRISES, Amar Infosoft Private Limited,"
        company_cell.font = Font(bold=True, size=10)
        company_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:S2')
        run_date_cell = ws['A2']
        run_date_cell.value = f"Run Date & Time: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        run_date_cell.font = Font(size=10)
        run_date_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A4:S4')
        title_cell = ws['A4']
        title_cell.value = "EMPLOYEE MASTER"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        headers = [
            'Srl. No.',
            'Pay Code',
            'Card No.',
            'Sex',
            'Birth Date',
            'Name',
            'Guardian\'s Name',
            'Department',
            'Section',
            'Desig/Category',
            'Joining Date',
            'Auth Shift',
            'Shift Type',
            'Shift Pattern',
            '1st Weekly Off',
            '2nd Weekly Off',
            '2nd Weekly Off',
            'OT Allowed/Rate',
            'Round the Clock'
        ]
        header_font = Font(bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[6].height = 30
        
        for row_num, employee in enumerate(queryset, 7):
            birth_date = employee.birth_date.strftime('%d/%m/%Y') if employee.birth_date else ''
            joining_date = employee.joining_date.strftime('%d/%m/%Y') if employee.joining_date else ''
            
            row_data = [
                row_num - 6,  # Serial number starting from 1
                employee.pay_code,
                employee.card_no,
                employee.sex,
                birth_date,
                employee.name.upper() if employee.name else '',
                employee.guardian_name.upper() if employee.guardian_name else '',
                employee.department,
                employee.section,
                employee.desig_category or '',
                joining_date,
                employee.auth_shift,
                employee.shift_type,
                employee.shift_pattern,
                employee.first_weekly_off,
                employee.second_weekly_off or '',
                employee.second_weekly_off_fh or '',
                'Y' if employee.ot_allowed_rate else 'N',
                'Y' if employee.round_the_clock else 'N'
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center" if col_num in [1, 3, 4, 18, 19] else "left", 
                    vertical="center"
                )
                cell.font = Font(size=9)
        
        # Set column widths
        column_widths = {
            'A': 6, 'B': 10, 'C': 10, 'D': 4, 'E': 12, 'F': 15, 'G': 15, 
            'H': 20, 'I': 20, 'J': 15, 'K': 12, 'L': 10, 'M': 10, 'N': 12, 
            'O': 12, 'P': 12, 'Q': 12, 'R': 10, 'S': 12
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'EMPLOYEE_MASTER_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response









 # management review dashboard
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import datetime
from .models import ManagementReview
from .serializers import (TrainingDataSerializer, DefectsDataSerializer,OperatorsChartSerializer, TrainingPlansChartSerializer,DefectsChartSerializer)

class CurrentMonthTrainingDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        try:
            data = ManagementReview.objects.get(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            serializer = TrainingDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class CurrentMonthDefectsDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        try:
            data = ManagementReview.objects.get(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            serializer = DefectsDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class OperatorsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = OperatorsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TrainingPlansChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = TrainingPlansChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DefectsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = DefectsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





# views.py

# import pandas as pd
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import AdvancedManpowerCTQ
# from .serializers import AdvancedManpowerCTQUploadSerializer

# class AdvancedManpowerCTQUploadView(APIView):
#     def post(self, request):
#         serializer = AdvancedManpowerCTQUploadSerializer(data=request.data)
#         if serializer.is_valid():
#             excel_file = serializer.validated_data['file']
#             try:
#                 # Read Excel with header at row 1
#                 df = pd.read_excel(excel_file, header=1)
#                 df.columns = df.columns.str.strip()  # Remove trailing/leading spaces

#                 for _, row in df.iterrows():
#                     AdvancedManpowerCTQ.objects.update_or_create(
#                         month_year_ctq=pd.to_datetime(row['Month & Year']).date(),
#                         defaults={
#                             'total_stations_ctq': int(row['Total Stations']),
#                             'operator_required_ctq': int(row['Operator Required']),
#                             'operator_availability_ctq': int(row['Operator Availability']),
#                             'buffer_manpower_required_ctq': int(row['Buffer Man Power Required']),
#                             'buffer_manpower_availability_ctq': int(row['Buffer Man Power Availability']),
#                             'attrition_trend_ctq': int(row['Attrition Trend']),
#                             'absentee_trend_ctq': int(row['Absentee Trend']),
#                             'planned_units_ctq': int(row['Planned Units']),
#                             'actual_production_ctq': int(row['Actual Production']),
#                         }
#                     )
#                 return Response({"message": "Data uploaded successfully."}, status=status.HTTP_201_CREATED)

#             except Exception as e:
#                 return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ManagementReview
from .serializers import ManagementReviewUploadSerializer

class ManagementReviewUploadView(APIView):
    def post(self, request):
        serializer = ManagementReviewUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            try:
                df = pd.read_excel(excel_file, header=1)
                df.columns = df.columns.str.strip()  # 🧽 Clean trailing/leading spaces from column names

                for _, row in df.iterrows():
                    ManagementReview.objects.update_or_create(
                        month_year=pd.to_datetime(row['Month & Year']).date(),
                        defaults={
                            'new_operators_joined': int(row['New Operators Joined']),
                            'new_operators_trained': int(row['New Operators Trained']),
                            'total_training_plans': int(row['Total Training Plans']),
                            'total_trainings_actual': int(row['Total Trainings Actual']),
                            'total_defects_msil': int(row['Total Defects at MSIL']),
                            'ctq_defects_msil': int(row['CTQ Defects at MSIL']),
                            'total_defects_tier1': int(row['Total Defects at Tier-1']),
                            'ctq_defects_tier1': int(row['CTQ Defects at Tier-1']),  # fixed trailing space issue
                            'total_internal_rejection': int(row['Total Internal Rejection']),
                            'ctq_internal_rejection': int(row['CTQ Internal Rejection']),
                        }
                    )
                return Response({"message": "Data uploaded successfully."}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from rest_framework import viewsets
from .models import ManagementReview
from .serializers import ManagementReviewSerializer

class ManagementReviewViewSet(viewsets.ModelViewSet):
    queryset = ManagementReview.objects.all().order_by('-month_year')
    serializer_class = ManagementReviewSerializer

from rest_framework import viewsets
from .models import CompanyLogo
from .serializers import CompanyLogoSerializer

class CompanyLogoViewSet(viewsets.ModelViewSet):
    queryset = CompanyLogo.objects.all()
    serializer_class = CompanyLogoSerializer



from rest_framework import generics
from .models import SubLine
from .serializers import SubLineSerializer

class SubLineListAPIView(generics.ListAPIView):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer




import pandas as pd
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import (
    OperatorSkill, EmployeeMaster, Station,
    SubLine, MainLine, MainDepartment
)

@csrf_exempt
def upload_operator_skills(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)

            failed_rows = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        pay_code = str(row['pay_code']).strip()
                        department_name = str(row['department']).strip()
                        main_line_name = str(row['main_line']).strip()
                        sub_line_name = str(row['sub_line']).strip()
                        skill_name = str(row['skill']).strip()
                        skill_level = str(row['skill_level']).strip()

                        # Get Employee
                        employee = EmployeeMaster.objects.get(pay_code=pay_code)

                        # Get Station via nested lookups
                        department = MainDepartment.objects.get(name=department_name)
                        main_line = MainLine.objects.get(name=main_line_name, department=department)
                        sub_line = SubLine.objects.get(name=sub_line_name, main_line=main_line)
                        station = Station.objects.get(skill=skill_name, sub_line=sub_line)

                        # Create or Update OperatorSkill
                        OperatorSkill.objects.update_or_create(
                            operator=employee,
                            station=station,
                            defaults={
                                'skill_level': skill_level
                            }
                        )

                    except Exception as e:
                        failed_rows.append({
                            'row_index': index + 2,
                            'error': str(e)
                        })
                        # Trigger rollback if any row fails
                        raise

            return JsonResponse({
                'status': 'success',
                'message': f"{df.shape[0]} records uploaded successfully.",
                'failed': []
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': 'Upload failed. No records were saved.',
                'error': str(e),
                'failed': failed_rows
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request. Upload a file using POST method.'})











class StationDeleteView(APIView):
    def delete(self, request, pk):
        station = get_object_or_404(Station, pk=pk)
        station.delete()
        return Response({"message": "Station deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

class StationUpdateView(APIView):
    def put(self, request, pk):
        station = get_object_or_404(Station, pk=pk)
        serializer = StationSerializer(station, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)





from .serializers import EmployeeWithStatusSerializer

class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = ...  # your main MachineAllocation serializer

    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        machine_id = request.query_params.get('machine_id')
        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=400)

        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({'error': 'Machine not found'}, status=404)

        matching_skills = OperatorSkill.objects.filter(station__skill=machine.process)
        employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
        employees = EmployeeMaster.objects.filter(id__in=employee_ids)

        serializer = EmployeeWithStatusSerializer(employees, many=True, context={'machine_id': machine_id})
        return Response(serializer.data)
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Department
from .serializers import DepartmentSerializer

class DepartmentByFactoryView(APIView):
    def get(self, request):
        factory_id = request.query_params.get('factory')
        if not factory_id:
            return Response({"error": "Factory ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        departments = Department.objects.filter(factory_id=factory_id)
        serializer = DepartmentSerializer(departments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework import viewsets
from .models import AdvancedManpowerCTQ
from .serializers import NewAdvancedManpowerCTQSerializer

class NewAdvancedManpowerCTQViewSet(viewsets.ModelViewSet):
    queryset = AdvancedManpowerCTQ.objects.all().order_by('-month_year_ctq')
    serializer_class = NewAdvancedManpowerCTQSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime
from .models import AdvancedManpowerCTQ, OperatorRequirement
from .serializers import (
    AdvancedManpowerCTQSerializer,
    OperatorTrendSerializer,
    BufferManpowerTrendSerializer,
    AttritionTrendSerializer,
    AbsenteeTrendSerializer,
    OperatorRequirementSerializer
)

class ManpowerCTQTrendsView(APIView):
    def get(self, request):
        plant = request.query_params.get('plant')
        factory_id = request.query_params.get('factory_id')
        department_id = request.query_params.get('department_id')
        today = datetime.today()

        # ===== CTQ Queryset =====
        ctq_queryset = AdvancedManpowerCTQ.objects.all().order_by('month_year_ctq')
        if plant:
            ctq_queryset = ctq_queryset.filter(plant=plant)
        if factory_id:
            ctq_queryset = ctq_queryset.filter(factory_id=factory_id)
        if department_id:
            ctq_queryset = ctq_queryset.filter(department_id=department_id)

        # Current month data
        current_month_ctq = ctq_queryset.filter(
            month_year_ctq__year=today.year,
            month_year_ctq__month=today.month
        )

        # ===== OperatorRequirement Queryset =====
        operator_queryset = OperatorRequirement.objects.all().order_by('-month')
        if factory_id:
            operator_queryset = operator_queryset.filter(factory_id=factory_id)
        if department_id:
            operator_queryset = operator_queryset.filter(department_id=department_id)

        # ===== Build Flat Response =====
        return Response({
            "current_month": AdvancedManpowerCTQSerializer(current_month_ctq, many=True).data,
            "operator_trend": OperatorTrendSerializer(ctq_queryset, many=True).data,
            "buffer_trend": BufferManpowerTrendSerializer(ctq_queryset, many=True).data,
            "attrition_trend": AttritionTrendSerializer(ctq_queryset, many=True).data,
            "absentee_trend": AbsenteeTrendSerializer(ctq_queryset, many=True).data,
            "operator_requirements": OperatorRequirementSerializer(operator_queryset, many=True).data,
        })
    
from rest_framework import viewsets
from .models import OperatorRequirement
from .serializers import OperatorRequirementSerializer

class OperatorRequirementViewSet(viewsets.ModelViewSet):
    queryset = OperatorRequirement.objects.all().order_by('-month')
    serializer_class = OperatorRequirementSerializer


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import AdvancedManpowerCTQ, Factory, Department, HQ
from .serializers import AdvancedManpowerCTQSerializer

class UploadAdvancedManpowerCTQView(APIView):
    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(excel_file)

            for _, row in df.iterrows():
                factory_name = row['factory_name']
                department_name = row['department_name']

                # Create or get dummy HQ (if not already assigned)
                hq, _ = HQ.objects.get_or_create(name='Default HQ')

                factory, _ = Factory.objects.get_or_create(name=factory_name, defaults={'hq': hq})
                department, _ = Department.objects.get_or_create(name=department_name, factory=factory)

                # Direct creation using model (instead of serializer if name usage)
                AdvancedManpowerCTQ.objects.create(
                    month_year_ctq=row['month_year_ctq'],
                    total_stations_ctq=row['total_stations_ctq'],
                    operator_required_ctq=row['operator_required_ctq'],
                    operator_availability_ctq=row['operator_availability_ctq'],
                    buffer_manpower_required_ctq=row['buffer_manpower_required_ctq'],
                    buffer_manpower_availability_ctq=row['buffer_manpower_availability_ctq'],
                    attrition_trend_ctq=row['attrition_trend_ctq'],
                    absentee_trend_ctq=row['absentee_trend_ctq'],
                    planned_units_ctq=row['planned_units_ctq'],
                    actual_production_ctq=row['actual_production_ctq'],
                    factory=factory,
                    department=department
                )

            return Response({"message": "Excel data uploaded successfully."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)








import os
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings

class LogoView(APIView):
    def get(self, request):
        image_path = 'logos/nl_logo.jpg'
        full_url = request.build_absolute_uri(settings.MEDIA_URL + image_path)
        return Response({'logo_url': full_url})

# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import FileResponse
from .models import UploadedFile
from .serializers import UploadedFileSerializer
import os

class UploadedFileListView(APIView):
    def get(self, request):
        files = UploadedFile.objects.all().order_by('-uploaded_at')
        serializer = UploadedFileSerializer(files, many=True)
        return Response(serializer.data)

class FileDownloadView(APIView):
    def get(self, request, file_id):
        try:
            uploaded_file = UploadedFile.objects.get(id=file_id)
            file_path = uploaded_file.file.path
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
        except UploadedFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=404)









from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import LevelOneScoreSerializer

class LevelOneEmployeesView(APIView):
    def get(self, request):
        level_one_scores = Score.objects.filter(level="Level 1").select_related('employee')
        serializer = LevelOneScoreSerializer(level_one_scores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import LevelTwoGroupedEmployeeScoreSerializer
from collections import defaultdict

class GroupedScoreByEmployeeView(APIView):
    def get(self, request):
        scores = Score.objects.filter(level="Level 2").select_related('employee', 'skill')

        grouped_data = defaultdict(list)
        for score in scores:
            key = (score.employee.id, score.employee.name)
            grouped_data[key].append(score)

        result = []
        for (employee_id, employee_name), score_list in grouped_data.items():
            result.append({
                "employee_id": employee_id,
                "employee_name": employee_name,
                "scores": score_list
            })

        serializer = LevelTwoGroupedEmployeeScoreSerializer(result, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from collections import defaultdict
from .models import Score
from .serializers import LevelThreeGroupedEmployeeScoreSerializer

class GroupedLevelThreeScoreByEmployeeView(APIView):
    def get(self, request):
        scores = Score.objects.filter(level="Level 3").select_related('employee', 'skill')

        grouped_data = defaultdict(list)
        for score in scores:
            key = (score.employee.id, score.employee.name)
            grouped_data[key].append(score)

        result = []
        for (employee_id, employee_name), score_list in grouped_data.items():
            result.append({
                "employee_id": employee_id,
                "employee_name": employee_name,
                "scores": score_list
            })

        serializer = LevelThreeGroupedEmployeeScoreSerializer(result, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)











# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from django_celery_beat.models import PeriodicTask, CrontabSchedule
# import json
# from datetime import datetime
# from django.utils import timezone
# import pytz

# class SetTaskTimeView(APIView):
#     def post(self, request):
#         time_str = request.data.get("time")  # Expecting format like "02:30 PM"

#         if not time_str:
#             return Response({"error": "Time is required (e.g., 02:30 PM)."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             # Convert 12-hour format with AM/PM to 24-hour format
#             time_obj = datetime.strptime(time_str.strip(), "%I:%M %p")
#             hour = time_obj.hour
#             minute = time_obj.minute
#         except ValueError:
#             return Response({"error": "Invalid time format. Use HH:MM AM/PM."}, status=status.HTTP_400_BAD_REQUEST)

#         # Delete old schedules for this task to avoid conflicts
#         old_task = PeriodicTask.objects.filter(name="daily_import_excel").first()
#         if old_task and old_task.crontab:
#             old_crontab = old_task.crontab
#             old_task.delete()
#             # Check if any other tasks use this schedule
#             if not PeriodicTask.objects.filter(crontab=old_crontab).exists():
#                 old_crontab.delete()

#         # Create new schedule
#         schedule, created = CrontabSchedule.objects.get_or_create(
#             minute=str(minute),
#             hour=str(hour),
#             day_of_week='*',
#             day_of_month='*',
#             month_of_year='*',
#         )

#         # Create new task
#         task = PeriodicTask.objects.create(
#             name="daily_import_excel",
#             crontab=schedule,
#             task='app1.tasks.import_attendance_from_excel',
#             enabled=True,
#             args=json.dumps([]),
#         )

#         # Get current time in Asia/Kolkata timezone
#         kolkata_tz = pytz.timezone('Asia/Kolkata')
#         current_kolkata_time = timezone.now().astimezone(kolkata_tz)

#         return Response({
#             "success": True,
#             "message": f"Task scheduled daily at {time_str} (24-hour: {hour:02}:{minute:02})",
#             "scheduled_time": f"{hour:02}:{minute:02}",
#             "current_time_utc": timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
#             "current_time_kolkata": current_kolkata_time.strftime('%Y-%m-%d %H:%M:%S'),
#             "timezone": "Asia/Kolkata",
#             "task_enabled": task.enabled
#         })
    



def schedule_task(task_name, task_path, time_str):
    from datetime import datetime
    from django.utils import timezone
    import pytz
    from django_celery_beat.models import PeriodicTask, CrontabSchedule
    import json

    try:
        dt = datetime.strptime(time_str.strip(), "%I:%M %p")
        hour, minute = dt.hour, dt.minute
    except ValueError:
        return None, "Invalid time format. Use HH:MM AM/PM."

    existing_task = PeriodicTask.objects.filter(name=task_name).first()
    if existing_task:
        old_cron = existing_task.crontab
        existing_task.delete()
        if not PeriodicTask.objects.filter(crontab=old_cron).exists():
            old_cron.delete()

    cron, _ = CrontabSchedule.objects.get_or_create(
        minute=str(minute), hour=str(hour),
        day_of_week='*', day_of_month='*', month_of_year='*',
    )

    task = PeriodicTask.objects.create(
        name=task_name,
        crontab=cron,
        task=task_path,
        args=json.dumps([]),
        enabled=True
    )

    kolkata = timezone.now().astimezone(pytz.timezone('Asia/Kolkata'))
    return {
        "message": f"{task_name} scheduled at {time_str} (24-hour: {hour:02}:{minute:02})",
        "current_kolkata_time": kolkata.strftime('%Y-%m-%d %H:%M:%S'),
        "task_enabled": task.enabled
    }, None





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class SetAttendanceTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_excel", "app1.tasks.import_attendance_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)

class SetManagementReviewTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_management_review", "app1.tasks.import_management_review_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)

class SetAdvancedManpowerTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_advanced_manpower", "app1.tasks.import_advanced_manpower_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)












# import pandas as pd
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import EmployeeMaster, Station, OperatorSkill


# from django.db.models import Max

# class UploadOperatorSkillsAPIView(APIView):
#     def post(self, request):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             df = pd.read_excel(file)

#             fixed_columns = ["S.No.", "Operator Name", "Code", "Date of Joining"]
#             station_columns = [col for col in df.columns if col not in fixed_columns]

#             skipped_employees = []
#             skipped_stations = []

#             from django.db.models import Func

#             class Trim(Func):
#                 function = 'TRIM'

#             # --- Step 1: Remove existing duplicates ---
#             duplicates = (
#                 OperatorSkill.objects
#                 .values('operator_id', 'station_id')
#                 .annotate(latest_id=Max('id'))
#             )
#             for dup in duplicates:
#                 OperatorSkill.objects.filter(
#                     operator_id=dup['operator_id'],
#                     station_id=dup['station_id']
#                 ).exclude(id=dup['latest_id']).delete()

#             # --- Step 2: Process the file ---
#             for _, row in df.iterrows():
#                 operators = EmployeeMaster.objects.annotate(
#                     trimmed_name=Trim('name')
#                 ).filter(trimmed_name__iexact=str(row['Operator Name']).strip())

#                 if not operators.exists():
#                     skipped_employees.append(str(row['Operator Name']))
#                     continue

#                 for operator in operators:
#                     for station_name in station_columns:
#                         skill_level = row.get(station_name)
#                         if pd.isna(skill_level) or skill_level == '':
#                             continue

#                         stations = Station.objects.filter(skill__iexact=station_name.strip())
#                         if not stations.exists():
#                             skipped_stations.append(station_name)
#                             continue

#                         for station in stations:
#                             OperatorSkill.objects.update_or_create(
#                                 operator=operator,
#                                 station=station,
#                                 defaults={'skill_level': str(skill_level).strip()}
#                             )

#             return Response({
#                 "message": "Skills updated successfully!",
#                 "skipped_employees": skipped_employees,
#                 "skipped_stations": list(set(skipped_stations))
#             }, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Max
from .models import EmployeeMaster, Station, OperatorSkill


class UploadOperatorSkillsAPIView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)

            fixed_columns = ["S.No.", "Operator Name", "Code", "Date of Joining"]
            station_columns = [col for col in df.columns if col not in fixed_columns]

            skipped_employees = []
            skipped_stations = []

            # --- Step 1: Remove existing duplicates ---
            duplicates = (
                OperatorSkill.objects
                .values('operator_id', 'station_id')
                .annotate(latest_id=Max('id'))
            )
            for dup in duplicates:
                OperatorSkill.objects.filter(
                    operator_id=dup['operator_id'],
                    station_id=dup['station_id']
                ).exclude(id=dup['latest_id']).delete()

            # --- Step 2: Process the file ---
            for _, row in df.iterrows():
                # Normalize Excel code (remove leading zeros)
                pay_code_excel = str(row['Code']).strip().lstrip("0")

                # Also normalize DB pay_code (remove leading zeros before comparing)
                operator = None
                for emp in EmployeeMaster.objects.all():
                    if emp.pay_code.strip().lstrip("0") == pay_code_excel:
                        operator = emp
                        break

                if not operator:
                    skipped_employees.append(pay_code_excel)
                    continue

                # Process station skill levels
                for station_name in station_columns:
                    skill_level = row.get(station_name)
                    if pd.isna(skill_level) or skill_level == '':
                        continue

                    stations = Station.objects.filter(skill__iexact=station_name.strip())
                    if not stations.exists():
                        skipped_stations.append(station_name)
                        continue

                    for station in stations:
                        OperatorSkill.objects.update_or_create(
                            operator=operator,
                            station=station,
                            defaults={'skill_level': str(skill_level).strip()}
                        )

            return Response({
                "message": "Skills updated successfully!",
                "skipped_employees": skipped_employees,
                "skipped_stations": list(set(skipped_stations))
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)